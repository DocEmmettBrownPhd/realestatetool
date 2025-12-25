from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def generate_pdf_report(data):
    filename = f"reports/analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    doc = SimpleDocTemplate(filename, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    story = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=24,
        spaceAfter=20,
        textColor=colors.HexColor('#1e293b')
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        spaceBefore=20,
        spaceAfter=10,
        textColor=colors.HexColor('#334155')
    )
    
    story.append(Paragraph("Real Estate Investment Analysis", title_style))
    story.append(Spacer(1, 12))
    story.append(Paragraph(f"<b>Property:</b> {data.get('address', 'N/A')}", styles['Normal']))
    story.append(Paragraph(f"<b>Generated:</b> {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Comps Summary
    if data.get('comps'):
        comps = data['comps']
        story.append(Paragraph("Market Analysis Summary", heading_style))
        
        summary_data = [
            ['Metric', 'Value'],
            ['Comparable Sales Found', str(comps.get('total_found', 0))],
            ['Average Sale Price', f"${comps.get('average_price', 0):,}"],
            ['Average Price/Sq Ft', f"${comps.get('average_price_per_sqft', 0):,.2f}"],
            ['Estimated ARV', f"${comps.get('estimated_value', 0):,}"]
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2.5*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ]))
        story.append(summary_table)
    
    # Investment Scenarios
    if data.get('scenarios'):
        story.append(Spacer(1, 20))
        story.append(Paragraph("Investment Scenarios", heading_style))
        story.append(Spacer(1, 12))
        
        table_data = [['Strategy', 'Investment', 'Profit', 'ROI', 'Timeline']]
        for scenario in data['scenarios']:
            roi = scenario['roi']
            roi_display = f"{roi}%" if roi > 0 else f"({abs(roi)}%)"
            timeline = f"{scenario.get('timeline_days', 0)} days"
            table_data.append([
                scenario['name'],
                f"${scenario['total_investment']:,}",
                f"${scenario.get('profit', 0):,}",
                roi_display,
                timeline
            ])
        
        table = Table(table_data, colWidths=[2*inch, 1.3*inch, 1.3*inch, 0.9*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#dcfce7')),
            ('BACKGROUND', (0, 2), (-1, -1), colors.HexColor('#f8fafc')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ]))
        story.append(table)
    
    # Best Strategy Highlight
    if data.get('best_scenario'):
        story.append(Spacer(1, 25))
        best = data['best_scenario']
        story.append(Paragraph(
            f"<b>Recommended Strategy:</b> {best['name']} with {best['roi']}% ROI",
            styles['Normal']
        ))
    
    doc.build(story)
    return filename

def generate_excel_report(data):
    filename = f"reports/analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb = openpyxl.Workbook()
    
    # Styles
    header_font = Font(size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    accent_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    # Summary Sheet
    ws = wb.active
    ws.title = "Summary"
    
    ws['A1'] = "Real Estate Investment Analysis"
    ws['A1'].font = Font(size=18, bold=True, color='0F172A')
    ws.merge_cells('A1:D1')
    
    ws['A2'] = f"Property: {data.get('address', 'N/A')}"
    ws['A3'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    
    # Scenarios Table
    ws['A5'] = "Investment Scenarios"
    ws['A5'].font = Font(size=14, bold=True)
    
    headers = ['Strategy', 'Investment', 'Profit', 'ROI', 'Timeline']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    if data.get('scenarios'):
        for idx, scenario in enumerate(data['scenarios'], start=7):
            ws.cell(row=idx, column=1, value=scenario['name']).border = border
            ws.cell(row=idx, column=2, value=scenario['total_investment']).border = border
            ws.cell(row=idx, column=2).number_format = '$#,##0'
            ws.cell(row=idx, column=3, value=scenario.get('profit', 0)).border = border
            ws.cell(row=idx, column=3).number_format = '$#,##0'
            ws.cell(row=idx, column=4, value=f"{scenario['roi']}%").border = border
            ws.cell(row=idx, column=5, value=f"{scenario.get('timeline_days', 0)} days").border = border
            
            # Highlight best scenario
            if idx == 7:
                for col in range(1, 6):
                    ws.cell(row=idx, column=col).fill = accent_fill
    
    # Column widths
    column_widths = [25, 15, 15, 12, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Comps Sheet
    if data.get('comps') and data['comps'].get('properties'):
        ws_comps = wb.create_sheet("Comparable Sales")
        
        ws_comps['A1'] = "Comparable Sales Analysis"
        ws_comps['A1'].font = Font(size=14, bold=True)
        
        comp_headers = ['Address', 'Sale Price', 'Beds', 'Baths', 'Sq Ft', '$/Sq Ft', 'Distance']
        for col, header in enumerate(comp_headers, 1):
            cell = ws_comps.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for idx, comp in enumerate(data['comps']['properties'], start=4):
            ws_comps.cell(row=idx, column=1, value=comp.get('address', 'N/A')).border = border
            ws_comps.cell(row=idx, column=2, value=comp.get('price', 0)).border = border
            ws_comps.cell(row=idx, column=2).number_format = '$#,##0'
            ws_comps.cell(row=idx, column=3, value=comp.get('beds', 0)).border = border
            ws_comps.cell(row=idx, column=4, value=comp.get('baths', 0)).border = border
            ws_comps.cell(row=idx, column=5, value=comp.get('sqft', 0)).border = border
            ws_comps.cell(row=idx, column=5).number_format = '#,##0'
            ws_comps.cell(row=idx, column=6, value=comp.get('price_per_sqft', 0)).border = border
            ws_comps.cell(row=idx, column=6).number_format = '$#,##0.00'
            ws_comps.cell(row=idx, column=7, value=f"{comp.get('distance_miles', 'N/A')} mi").border = border
        
        comp_widths = [35, 15, 8, 8, 12, 12, 12]
        for i, width in enumerate(comp_widths, 1):
            ws_comps.column_dimensions[get_column_letter(i)].width = width
    
    wb.save(filename)
    return filename
