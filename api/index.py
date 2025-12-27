from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS
import requests
import os
import json
import time
from datetime import datetime
from io import BytesIO

app = Flask(__name__, static_folder='../public')
CORS(app)

APIFY_TOKEN = os.getenv('APIFY_API_TOKEN', 'apify_api_CHtm8I3iS00QsiRaNozGNMQppjZuGJ2sp0cp')

def haversine_distance(lat1, lon1, lat2, lon2):
    from math import radians, sin, cos, sqrt, atan2
    R = 3959
    lat1, lon1, lat2, lon2 = map(radians, [lat1, lon1, lat2, lon2])
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
    c = 2 * atan2(sqrt(a), sqrt(1-a))
    return R * c

def fetch_subject_property(address):
    if not APIFY_TOKEN:
        return None
    
    actor_input = {"addresses": address}
    
    try:
        # Use synchronous call
        response = requests.post(
            'https://api.apify.com/v2/acts/aknahin~zillow-property-info-scraper/run-sync-get-dataset-items?token=' + APIFY_TOKEN,
            json=actor_input,
            timeout=300
        )
        
        if response.status_code != 200 and response.status_code != 201:
            return None
        
        properties = response.json()
        
        if properties and len(properties) > 0:
            prop = properties[0]
            if prop.get('error'):
                return None
            
            full_address = prop.get('address', address)
            address_parts = full_address.split(',')
            city = address_parts[1].strip() if len(address_parts) > 1 else ''
            state_zip = address_parts[2].strip() if len(address_parts) > 2 else ''
            state = state_zip.split()[0] if state_zip else ''
            zipcode = state_zip.split()[1] if len(state_zip.split()) > 1 else ''
            
            return {
                'address': full_address,
                'city': city,
                'state': state,
                'zipcode': zipcode,
                'beds': prop.get('beds', 3),
                'baths': prop.get('baths', 2),
                'sqft': prop.get('area', 1800),
                'year_built': prop.get('yearBuilt', 2000),
                'lot_size': prop.get('lotSize', 0.25),
                'latitude': prop.get('latLong', {}).get('latitude'),
                'longitude': prop.get('latLong', {}).get('longitude'),
                'zestimate': prop.get('zestimate', 0),
                'zpid': prop.get('zpid', ''),
                'status': prop.get('statusText', 'Unknown'),
                'image_url': prop.get('imgSrc', '')
            }
        return None
    except Exception as e:
        return None

def scrape_zillow_comps(zipcode, beds, baths, sqft, year_built):
    if not APIFY_TOKEN:
        return get_demo_comps(zipcode, sqft)
    
    min_year = max(1900, year_built - 10) if year_built else 1900
    max_year = (year_built + 10) if year_built else 2025
    min_sqft_range = str(int(sqft * 0.8))
    max_sqft_range = str(int(sqft * 1.2))
    
    actor_input = {
        "location": zipcode,
        "operation": "sold",
        "sortBy": "newest",
        "minBeds": max(1, beds - 1),
        "maxBeds": beds + 1,
        "minBaths": max(1, baths - 1),
        "homeTypes": ["houses"],
        "minYearBuilt": min_year,
        "maxYearBuilt": max_year,
        "minSize": min_sqft_range,
        "maxSize": max_sqft_range,
        "maxSoldDate": "6m",
        "maxItems": 20
    }
    
    try:
        # Use synchronous call with waitForFinish
        response = requests.post(
            'https://api.apify.com/v2/acts/igolaizola~zillow-scraper-ppe/run-sync-get-dataset-items?token=' + APIFY_TOKEN,
            json=actor_input,
            timeout=300
        )
        
        if response.status_code != 200 and response.status_code != 201:
            return get_demo_comps(zipcode, sqft)
        
        comps = response.json()
        processed_comps = []
        
        for comp in comps:
            if (comp.get('bedrooms') and comp.get('bathrooms') and 
                comp.get('livingArea') and comp.get('price')):
                
                if not comp.get('address'):
                    comp['address'] = {
                        'streetAddress': comp.get('streetAddress', 'Unknown'),
                        'city': comp.get('city', ''),
                        'state': comp.get('state', 'GA'),
                        'zipcode': comp.get('zipcode', zipcode)
                    }
                if not isinstance(comp.get('price'), dict):
                    comp['price'] = {'value': comp.get('price', 0)}
                
                comp['price_per_sqft'] = round(comp['price']['value'] / comp['livingArea'], 2)
                comp['distance_miles'] = 0.0
                comp['latitude'] = comp.get('latitude', comp.get('latLong', {}).get('latitude'))
                comp['longitude'] = comp.get('longitude', comp.get('latLong', {}).get('longitude'))
                processed_comps.append(comp)
        
        if processed_comps:
            return processed_comps[:10]
        return get_demo_comps(zipcode, sqft)
    except Exception as e:
        return get_demo_comps(zipcode, sqft)

def calculate_distances(comps, subject_lat, subject_lon):
    if not subject_lat or not subject_lon:
        return comps
    
    for comp in comps:
        if comp.get('latitude') and comp.get('longitude'):
            distance = haversine_distance(subject_lat, subject_lon, comp['latitude'], comp['longitude'])
            comp['distance_miles'] = round(distance, 2)
        else:
            comp['distance_miles'] = 999
    
    comps.sort(key=lambda x: x['distance_miles'])
    filtered = [c for c in comps if c['distance_miles'] <= 2.0]
    return filtered if filtered else comps[:5]

def get_demo_comps(zipcode, sqft):
    base_price = 150
    comps = []
    for i in range(5):
        price_variation = base_price + (i * 5 - 10)
        comp_sqft = sqft + (i * 100 - 200)
        price = int(price_variation * comp_sqft)
        comps.append({
            'address': {'streetAddress': f'{1000 + i} Demo Street', 'city': 'Atlanta', 'state': 'GA', 'zipcode': zipcode},
            'price': {'value': price},
            'bedrooms': 3,
            'bathrooms': 2,
            'livingArea': comp_sqft,
            'yearBuilt': 2000,
            'price_per_sqft': round(price / comp_sqft, 2),
            'distance_miles': round(0.3 + i * 0.1, 2),
            'listing': {'dateSold': '2024-12-01'}
        })
    return comps

# Atlanta Metro FMR 2024
FMR_RATES = {
    '30002': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30004': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30005': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30008': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30012': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30013': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30014': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30016': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30017': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30019': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30021': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30024': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30030': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30032': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30033': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30034': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30035': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30038': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30039': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30040': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30041': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30043': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30044': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30045': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30046': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30047': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30052': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30058': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30060': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30062': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30064': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30066': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30067': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30068': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30071': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30072': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30075': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30076': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30078': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30079': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30080': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30082': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30083': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30084': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30087': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30088': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30092': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30093': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30094': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30096': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30097': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30101': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30102': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30106': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30107': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30114': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30115': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30126': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30127': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30134': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30135': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30137': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30141': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30144': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30152': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30157': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30168': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30180': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30188': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30189': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30213': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30214': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30215': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30228': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30236': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30238': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30248': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30252': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30253': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30260': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30269': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30273': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30274': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30281': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30288': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30291': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30294': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30296': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30297': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30301': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30303': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30305': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30306': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30307': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30308': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30309': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30310': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30311': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30312': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30313': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30314': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30315': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30316': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30317': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30318': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30319': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30322': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30324': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30326': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30327': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30328': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30329': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30331': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30332': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30334': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30336': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30337': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30338': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30339': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30340': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30341': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30342': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30344': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30345': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30349': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30350': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30354': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30360': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    '30363': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169},
    'default': {'0br': 1089, '1br': 1199, '2br': 1409, '3br': 1829, '4br': 2169}
}

def get_fmr(zipcode, beds):
    fmr_data = FMR_RATES.get(zipcode, FMR_RATES['default'])
    bed_key = f'{min(beds, 4)}br' if beds > 0 else '0br'
    return fmr_data.get(bed_key, fmr_data['2br'])

def calculate_rental_scenarios(property_data, estimated_arv):
    purchase = property_data['purchasePrice']
    sqft = property_data['currentSqft']
    beds = property_data['beds']
    baths = property_data['baths']
    zipcode = property_data.get('zipcode', '30344')
    
    # Buy & Hold Financing: 20% down, 7% rate, 30 year
    down_payment_pct = 0.20
    down_payment = purchase * down_payment_pct
    loan_amount = purchase - down_payment
    interest_rate = 0.07
    loan_term_years = 30
    closing_costs = purchase * 0.03
    
    # Monthly mortgage (P&I)
    monthly_rate = interest_rate / 12
    num_payments = loan_term_years * 12
    if monthly_rate > 0:
        monthly_mortgage = loan_amount * (monthly_rate * (1 + monthly_rate)**num_payments) / ((1 + monthly_rate)**num_payments - 1)
    else:
        monthly_mortgage = loan_amount / num_payments
    
    cash_invested = down_payment + closing_costs
    
    rental_scenarios = []
    
    # 1. Open Market Rental
    open_market_rent = sqft * 0.85
    open_market_rent = max(open_market_rent, get_fmr(zipcode, beds) * 0.9)
    
    # 50% Rule expenses breakdown (target 1.5% rule = rent >= 1.5% of purchase)
    vacancy_rate = 0.08  # 8%
    mgmt_rate = 0.10  # 10%
    repairs_rate = 0.08  # 8%
    capex_rate = 0.08  # 8%
    taxes_rate = 0.10  # 10% of rent (or ~1.2% of value)
    insurance_rate = 0.06  # 6%
    # Total: 50%
    
    gross_rent = open_market_rent
    vacancy = gross_rent * vacancy_rate
    egi = gross_rent - vacancy
    
    mgmt = gross_rent * mgmt_rate
    repairs = gross_rent * repairs_rate
    capex = gross_rent * capex_rate
    taxes = gross_rent * taxes_rate
    insurance = gross_rent * insurance_rate
    total_expenses = mgmt + repairs + capex + taxes + insurance
    
    noi = egi - total_expenses
    cash_flow = noi - monthly_mortgage
    annual_cash_flow = cash_flow * 12
    
    coc_return = (annual_cash_flow / cash_invested * 100) if cash_invested > 0 else 0
    cap_rate = ((noi * 12) / purchase * 100) if purchase > 0 else 0
    one_pct_rule = (gross_rent / purchase * 100) if purchase > 0 else 0
    dscr = (noi / monthly_mortgage) if monthly_mortgage > 0 else 0
    
    rental_scenarios.append({
        'name': 'Open Market Rental',
        'type': 'rental',
        'monthly_rent': round(gross_rent),
        'vacancy_rate': vacancy_rate * 100,
        'vacancy': round(vacancy),
        'effective_gross_income': round(egi),
        'expenses': {
            'management': round(mgmt),
            'repairs': round(repairs),
            'capex': round(capex),
            'taxes': round(taxes),
            'insurance': round(insurance),
            'total': round(total_expenses)
        },
        'expense_ratio': 50,
        'noi': round(noi),
        'financing': {
            'down_payment': round(down_payment),
            'down_payment_pct': down_payment_pct * 100,
            'loan_amount': round(loan_amount),
            'interest_rate': interest_rate * 100,
            'loan_term': loan_term_years,
            'closing_costs': round(closing_costs)
        },
        'monthly_mortgage': round(monthly_mortgage),
        'monthly_cash_flow': round(cash_flow),
        'annual_cash_flow': round(annual_cash_flow),
        'cash_invested': round(cash_invested),
        'cash_on_cash': round(coc_return, 1),
        'cap_rate': round(cap_rate, 1),
        'dscr': round(dscr, 2),
        'one_pct_rule': round(one_pct_rule, 2),
        'roi': round(coc_return, 1),
        'timeline': 'Ongoing',
        'risk_level': 'Medium',
        'management_intensity': 'Medium'
    })
    
    # 2. Section 8 Rental
    fmr = get_fmr(zipcode, beds)
    s8_rent = fmr * 1.0
    
    vacancy_rate_s8 = 0.05  # Lower vacancy for S8
    
    gross_rent_s8 = s8_rent
    vacancy_s8 = gross_rent_s8 * vacancy_rate_s8
    egi_s8 = gross_rent_s8 - vacancy_s8
    
    mgmt_s8 = gross_rent_s8 * mgmt_rate
    repairs_s8 = gross_rent_s8 * repairs_rate
    capex_s8 = gross_rent_s8 * capex_rate
    taxes_s8 = gross_rent_s8 * taxes_rate
    insurance_s8 = gross_rent_s8 * insurance_rate
    total_expenses_s8 = mgmt_s8 + repairs_s8 + capex_s8 + taxes_s8 + insurance_s8
    
    noi_s8 = egi_s8 - total_expenses_s8
    cash_flow_s8 = noi_s8 - monthly_mortgage
    annual_cash_flow_s8 = cash_flow_s8 * 12
    
    coc_return_s8 = (annual_cash_flow_s8 / cash_invested * 100) if cash_invested > 0 else 0
    cap_rate_s8 = ((noi_s8 * 12) / purchase * 100) if purchase > 0 else 0
    dscr_s8 = (noi_s8 / monthly_mortgage) if monthly_mortgage > 0 else 0
    
    rental_scenarios.append({
        'name': 'Section 8 Rental',
        'type': 'rental',
        'monthly_rent': round(s8_rent),
        'fmr': round(fmr),
        'vacancy_rate': vacancy_rate_s8 * 100,
        'vacancy': round(vacancy_s8),
        'effective_gross_income': round(egi_s8),
        'expenses': {
            'management': round(mgmt_s8),
            'repairs': round(repairs_s8),
            'capex': round(capex_s8),
            'taxes': round(taxes_s8),
            'insurance': round(insurance_s8),
            'total': round(total_expenses_s8)
        },
        'expense_ratio': 50,
        'noi': round(noi_s8),
        'financing': {
            'down_payment': round(down_payment),
            'down_payment_pct': down_payment_pct * 100,
            'loan_amount': round(loan_amount),
            'interest_rate': interest_rate * 100,
            'loan_term': loan_term_years,
            'closing_costs': round(closing_costs)
        },
        'monthly_mortgage': round(monthly_mortgage),
        'monthly_cash_flow': round(cash_flow_s8),
        'annual_cash_flow': round(annual_cash_flow_s8),
        'cash_invested': round(cash_invested),
        'cash_on_cash': round(coc_return_s8, 1),
        'cap_rate': round(cap_rate_s8, 1),
        'dscr': round(dscr_s8, 2),
        'roi': round(coc_return_s8, 1),
        'timeline': 'Ongoing',
        'risk_level': 'Low',
        'management_intensity': 'Low-Medium',
        'government_portion': '70%',
        'tenant_portion': '30%'
    })
    
    # 3. Rent-by-Room
    room_prices = []
    if beds >= 1:
        room_prices.append(650)
    if beds >= 2:
        room_prices.append(550)
    if beds >= 3:
        room_prices.append(500)
    if beds >= 4:
        room_prices.append(450)
    if beds >= 5:
        room_prices.append(400)
    
    total_room_rent = sum(room_prices)
    vacancy_rate_room = 0.10
    utilities_landlord = 150
    
    gross_rent_room = total_room_rent
    vacancy_room = gross_rent_room * vacancy_rate_room
    egi_room = gross_rent_room - vacancy_room
    
    mgmt_room = gross_rent_room * 0.12  # Higher mgmt
    repairs_room = gross_rent_room * 0.10  # Higher repairs
    capex_room = gross_rent_room * 0.08
    taxes_room = gross_rent_room * 0.10
    insurance_room = gross_rent_room * 0.06
    total_expenses_room = mgmt_room + repairs_room + capex_room + taxes_room + insurance_room + utilities_landlord
    
    noi_room = egi_room - total_expenses_room
    cash_flow_room = noi_room - monthly_mortgage
    annual_cash_flow_room = cash_flow_room * 12
    
    coc_return_room = (annual_cash_flow_room / cash_invested * 100) if cash_invested > 0 else 0
    cap_rate_room = ((noi_room * 12) / purchase * 100) if purchase > 0 else 0
    dscr_room = (noi_room / monthly_mortgage) if monthly_mortgage > 0 else 0
    
    rental_scenarios.append({
        'name': 'Rent-by-Room',
        'type': 'rental',
        'monthly_rent': round(total_room_rent),
        'room_breakdown': room_prices,
        'num_rooms': len(room_prices),
        'vacancy_rate': vacancy_rate_room * 100,
        'vacancy': round(vacancy_room),
        'effective_gross_income': round(egi_room),
        'expenses': {
            'management': round(mgmt_room),
            'repairs': round(repairs_room),
            'capex': round(capex_room),
            'taxes': round(taxes_room),
            'insurance': round(insurance_room),
            'utilities': round(utilities_landlord),
            'total': round(total_expenses_room)
        },
        'noi': round(noi_room),
        'financing': {
            'down_payment': round(down_payment),
            'down_payment_pct': down_payment_pct * 100,
            'loan_amount': round(loan_amount),
            'interest_rate': interest_rate * 100,
            'loan_term': loan_term_years,
            'closing_costs': round(closing_costs)
        },
        'monthly_mortgage': round(monthly_mortgage),
        'monthly_cash_flow': round(cash_flow_room),
        'annual_cash_flow': round(annual_cash_flow_room),
        'cash_invested': round(cash_invested),
        'cash_on_cash': round(coc_return_room, 1),
        'cap_rate': round(cap_rate_room, 1),
        'dscr': round(dscr_room, 2),
        'roi': round(coc_return_room, 1),
        'timeline': 'Ongoing',
        'risk_level': 'Medium-High',
        'management_intensity': 'High'
    })
    
    return rental_scenarios

def calculate_flip_scenarios(property_data, estimated_arv):
    purchase = property_data['purchasePrice']
    sqft = property_data['currentSqft']
    
    scenarios = []
    
    rehab_costs = {
        'light': sqft * 25,
        'medium': sqft * 45,
        'heavy': sqft * 75
    }
    
    arv_multipliers = {
        'light': 1.0,
        'medium': 1.05,
        'heavy': 1.15
    }
    
    for level, rehab in rehab_costs.items():
        holding_time = 4 if level == 'light' else 6 if level == 'medium' else 8
        
        # Hard Money: 10% down, 3 points, 10% interest
        hard_money_down_pct = 0.10
        hard_money_points = 0.03
        hard_money_rate = 0.10
        
        arv = estimated_arv * arv_multipliers[level]
        
        # 70% Rule check
        max_purchase_70 = (arv * 0.70) - rehab
        meets_70_rule = purchase <= max_purchase_70
        
        # Financing
        total_project_cost = purchase + rehab
        loan_amount = total_project_cost * (1 - hard_money_down_pct)
        down_payment = total_project_cost * hard_money_down_pct
        points_cost = loan_amount * hard_money_points
        interest = loan_amount * (hard_money_rate / 12) * holding_time
        
        # Costs
        closing_buy = purchase * 0.02
        closing_sell = arv * 0.06  # Agent commissions + closing
        holding_costs = holding_time * 500  # Utilities, insurance, etc.
        
        # Totals
        total_costs = purchase + rehab + points_cost + interest + closing_buy + closing_sell + holding_costs
        total_cash_needed = down_payment + points_cost + closing_buy + holding_costs
        
        # Profit calculations
        gross_profit = arv - purchase - rehab
        gross_profit_pct = (gross_profit / arv * 100) if arv > 0 else 0  # Target 30%
        
        net_profit = arv - total_costs
        net_profit_pct = (net_profit / arv * 100) if arv > 0 else 0  # Target 10%
        
        roi = (net_profit / total_cash_needed * 100) if total_cash_needed > 0 else 0
        
        scenarios.append({
            'name': f'Fix & Flip ({level.title()})',
            'type': 'flip',
            'purchase_price': round(purchase),
            'rehab_cost': round(rehab),
            'arv': round(arv),
            'max_purchase_70_rule': round(max_purchase_70),
            'meets_70_rule': meets_70_rule,
            'financing': {
                'down_payment': round(down_payment),
                'down_payment_pct': hard_money_down_pct * 100,
                'loan_amount': round(loan_amount),
                'points': round(points_cost),
                'points_pct': hard_money_points * 100,
                'interest_rate': hard_money_rate * 100,
                'interest_cost': round(interest),
                'holding_months': holding_time
            },
            'costs': {
                'purchase': round(purchase),
                'rehab': round(rehab),
                'points': round(points_cost),
                'interest': round(interest),
                'closing_buy': round(closing_buy),
                'closing_sell': round(closing_sell),
                'holding': round(holding_costs),
                'total': round(total_costs)
            },
            'total_investment': round(total_costs),
            'cash_needed': round(total_cash_needed),
            'sale_price': round(arv),
            'gross_profit': round(gross_profit),
            'gross_profit_pct': round(gross_profit_pct, 1),
            'profit': round(net_profit),
            'net_profit_pct': round(net_profit_pct, 1),
            'roi': round(roi, 1),
            'timeline': f'{holding_time} months',
            'timeline_days': holding_time * 30,
            'details': {
                'purchase': purchase,
                'rehab': round(rehab),
                'interest': round(interest),
                'points': round(points_cost),
                'closing_buy': round(closing_buy),
                'closing_sell': round(closing_sell),
                'holding': round(holding_costs),
                'arv': round(arv)
            }
        })
    
    # Wholesale
    assignment_fee = purchase * 0.06
    earnest_money = 1000
    wholesale_roi = (assignment_fee / earnest_money * 100)
    
    scenarios.append({
        'name': 'Wholesale Assignment',
        'type': 'wholesale',
        'purchase_price': round(purchase),
        'total_investment': earnest_money,
        'cash_needed': earnest_money,
        'profit': round(assignment_fee),
        'roi': round(wholesale_roi, 1),
        'timeline': '30 days',
        'timeline_days': 30,
        'details': {
            'contract_price': purchase,
            'assignment_fee': round(assignment_fee),
            'buyer_price': round(purchase + assignment_fee),
            'earnest_money': earnest_money
        }
    })
    
    return scenarios

@app.route('/api/lookup-property', methods=['POST'])
def lookup_property():
    try:
        data = request.json
        address = data.get('address', '')
        
        if not address:
            return jsonify({'error': 'Address required'}), 400
        
        property_details = fetch_subject_property(address)
        
        if property_details:
            return jsonify(property_details)
        else:
            return jsonify({'error': 'Property not found'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/analyze', methods=['POST'])
def analyze_property():
    try:
        data = request.json
        
        property_data = {
            'address': data['address'],
            'purchasePrice': float(data['purchasePrice']),
            'currentSqft': float(data['currentSqft']),
            'beds': int(data.get('beds', 3)),
            'baths': float(data.get('baths', 2)),
            'lotSize': float(data.get('lotSize', 0.25)),
            'zipcode': data.get('zipcode', data['address'].split()[-1] if data['address'] else '30344'),
            'yearBuilt': int(data.get('yearBuilt', 2000)),
            'latitude': data.get('latitude'),
            'longitude': data.get('longitude'),
            'zestimate': data.get('zestimate', 0)
        }
        
        comps = scrape_zillow_comps(
            property_data['zipcode'],
            property_data['beds'],
            property_data['baths'],
            property_data['currentSqft'],
            property_data['yearBuilt']
        )
        
        if property_data.get('latitude') and property_data.get('longitude'):
            comps = calculate_distances(comps, property_data['latitude'], property_data['longitude'])
        
        avg_price = sum(c['price']['value'] for c in comps) / len(comps) if comps else 0
        avg_price_per_sqft = sum(c['price_per_sqft'] for c in comps) / len(comps) if comps else 150
        estimated_arv = avg_price_per_sqft * property_data['currentSqft']
        
        # Get all scenarios
        flip_scenarios = calculate_flip_scenarios(property_data, estimated_arv)
        rental_scenarios = calculate_rental_scenarios(property_data, estimated_arv)
        
        all_scenarios = flip_scenarios + rental_scenarios
        all_scenarios.sort(key=lambda x: x['roi'], reverse=True)
        
        result = {
            'address': property_data['address'],
            'zestimate': property_data['zestimate'],
            'propertyData': property_data,
            'comps': {
                'total_found': len(comps),
                'average_price': round(avg_price),
                'average_price_per_sqft': round(avg_price_per_sqft, 2),
                'estimated_value': round(estimated_arv),
                'properties': comps[:5]
            },
            'scenarios': all_scenarios,
            'flip_scenarios': flip_scenarios,
            'rental_scenarios': rental_scenarios,
            'best_scenario': all_scenarios[0] if all_scenarios else None,
            'best_flip': flip_scenarios[0] if flip_scenarios else None,
            'best_rental': max(rental_scenarios, key=lambda x: x['roi']) if rental_scenarios else None
        }
        
        return jsonify(result)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/report/pdf', methods=['POST'])
def create_pdf_report():
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.enums import TA_CENTER
        
        data = request.json
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        story = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#1E293B'), spaceAfter=15, alignment=TA_CENTER)
        heading_style = ParagraphStyle('Heading', parent=styles['Heading2'], fontSize=14, textColor=colors.HexColor('#1E293B'), spaceAfter=10, spaceBefore=15)
        
        # Title
        story.append(Paragraph("Real Estate Investment Analysis", title_style))
        story.append(Paragraph(f"<b>Address:</b> {data.get('address', 'N/A')}", styles['Normal']))
        story.append(Paragraph(f"<b>Date:</b> {datetime.now().strftime('%B %d, %Y')}", styles['Normal']))
        
        prop = data.get('propertyData', {})
        story.append(Paragraph(f"<b>Purchase Price:</b> ${prop.get('purchasePrice', 0):,.0f} | <b>SqFt:</b> {prop.get('currentSqft', 0):,.0f} | <b>Beds/Baths:</b> {prop.get('beds', 0)}/{prop.get('baths', 0)}", styles['Normal']))
        story.append(Spacer(1, 15))
        
        # Comps Section
        if data.get('comps') and data['comps'].get('properties'):
            story.append(Paragraph("Comparable Sales", heading_style))
            
            # Summary
            comps = data['comps']
            summary_data = [
                ['Comps Found', 'Avg Price', 'Avg $/SqFt', 'Estimated ARV'],
                [str(comps.get('total_found', 0)), f"${comps.get('average_price', 0):,.0f}", f"${comps.get('average_price_per_sqft', 0):.2f}", f"${comps.get('estimated_value', 0):,.0f}"]
            ]
            t = Table(summary_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3B82F6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E2E8F0'))
            ]))
            story.append(t)
            story.append(Spacer(1, 10))
            
            # Individual comps
            comp_data = [['Address', 'Price', 'Beds/Bath', 'SqFt', '$/SqFt', 'Dist']]
            for comp in comps.get('properties', [])[:5]:
                addr = comp.get('address', {})
                street = addr.get('streetAddress', comp.get('streetAddress', 'N/A'))
                price = comp.get('price', {})
                price_val = price.get('value', price) if isinstance(price, dict) else price
                comp_data.append([
                    street[:25],
                    f"${price_val:,.0f}",
                    f"{comp.get('bedrooms', 0)}/{comp.get('bathrooms', 0)}",
                    f"{comp.get('livingArea', 0):,}",
                    f"${comp.get('price_per_sqft', 0):.0f}",
                    f"{comp.get('distance_miles', 0):.1f}mi"
                ])
            
            t = Table(comp_data, colWidths=[1.8*inch, 0.9*inch, 0.7*inch, 0.7*inch, 0.6*inch, 0.6*inch])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#64748B')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E2E8F0'))
            ]))
            story.append(t)
        
        story.append(Spacer(1, 15))
        
        # Flip Scenarios
        if data.get('flip_scenarios'):
            story.append(Paragraph("Fix & Flip Analysis (10% Down, 3 Points, 10% Interest)", heading_style))
            
            for s in data['flip_scenarios']:
                scenario_title = f"<b>{s['name']}</b>"
                if s.get('meets_70_rule') == True:
                    scenario_title += " <font color='green'>[MEETS 70% RULE]</font>"
                elif s.get('meets_70_rule') == False:
                    scenario_title += " <font color='red'>[FAILS 70% RULE]</font>"
                story.append(Paragraph(scenario_title, styles['Normal']))
                
                if s.get('costs'):
                    costs = s['costs']
                    flip_data = [
                        ['Purchase', 'Rehab', 'Points', 'Interest', 'Closing', 'Holding', 'Total Costs'],
                        [f"${costs.get('purchase', 0):,.0f}", f"${costs.get('rehab', 0):,.0f}", f"${costs.get('points', 0):,.0f}", 
                         f"${costs.get('interest', 0):,.0f}", f"${costs.get('closing_buy', 0) + costs.get('closing_sell', 0):,.0f}",
                         f"${costs.get('holding', 0):,.0f}", f"${costs.get('total', 0):,.0f}"]
                    ]
                    t = Table(flip_data, colWidths=[0.85*inch]*7)
                    t.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F59E0B')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), 7),
                        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E2E8F0'))
                    ]))
                    story.append(t)
                
                profit_data = [
                    ['ARV', 'Max (70%)', 'Cash Needed', 'Gross Profit', 'Net Profit', 'ROI', 'Timeline'],
                    [f"${s.get('arv', 0):,.0f}", f"${s.get('max_purchase_70_rule', 0):,.0f}", f"${s.get('cash_needed', 0):,.0f}",
                     f"${s.get('gross_profit', 0):,.0f} ({s.get('gross_profit_pct', 0)}%)", 
                     f"${s.get('profit', 0):,.0f} ({s.get('net_profit_pct', 0)}%)",
                     f"{s.get('roi', 0)}%", s.get('timeline', '')]
                ]
                t = Table(profit_data, colWidths=[0.85*inch]*7)
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#22C55E')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 7),
                    ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E2E8F0'))
                ]))
                story.append(t)
                story.append(Spacer(1, 8))
        
        story.append(PageBreak())
        
        # Rental Scenarios
        if data.get('rental_scenarios'):
            story.append(Paragraph("Rental Analysis (20% Down, 7% Rate, 50% Expense Rule)", heading_style))
            
            for s in data['rental_scenarios']:
                story.append(Paragraph(f"<b>{s['name']}</b>", styles['Normal']))
                
                # Income section
                income_data = [
                    ['Monthly Rent', 'Vacancy', 'EGI', 'Total Expenses', 'NOI', 'Mortgage', 'Cash Flow'],
                    [f"${s.get('monthly_rent', 0):,.0f}", f"${s.get('vacancy', 0):,.0f} ({s.get('vacancy_rate', 0)}%)", 
                     f"${s.get('effective_gross_income', 0):,.0f}", 
                     f"${s.get('expenses', {}).get('total', 0):,.0f}",
                     f"${s.get('noi', 0):,.0f}", f"${s.get('monthly_mortgage', 0):,.0f}", 
                     f"${s.get('monthly_cash_flow', 0):,.0f}"]
                ]
                t = Table(income_data, colWidths=[0.85*inch]*7)
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3B82F6')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 7),
                    ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E2E8F0'))
                ]))
                story.append(t)
                
                # Expenses breakdown
                if s.get('expenses'):
                    exp = s['expenses']
                    exp_data = [
                        ['Mgmt (10%)', 'Repairs (8%)', 'CapEx (8%)', 'Taxes (10%)', 'Insurance (6%)', 'Total (50%)'],
                        [f"${exp.get('management', 0):,.0f}", f"${exp.get('repairs', 0):,.0f}", f"${exp.get('capex', 0):,.0f}",
                         f"${exp.get('taxes', 0):,.0f}", f"${exp.get('insurance', 0):,.0f}", f"${exp.get('total', 0):,.0f}"]
                    ]
                    t = Table(exp_data, colWidths=[1*inch]*6)
                    t.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#64748B')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), 7),
                        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E2E8F0'))
                    ]))
                    story.append(t)
                
                # Returns
                returns_data = [
                    ['Cash Invested', 'Annual CF', 'Cash-on-Cash', 'Cap Rate', 'DSCR', '1% Rule'],
                    [f"${s.get('cash_invested', 0):,.0f}", f"${s.get('annual_cash_flow', 0):,.0f}",
                     f"{s.get('cash_on_cash', 0)}%", f"{s.get('cap_rate', 0)}%", 
                     f"{s.get('dscr', 0)}", f"{s.get('one_pct_rule', 0)}%"]
                ]
                t = Table(returns_data, colWidths=[1*inch]*6)
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#22C55E')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 7),
                    ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E2E8F0'))
                ]))
                story.append(t)
                story.append(Spacer(1, 10))
        
        doc.build(story)
        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name='analysis_report.pdf', mimetype='application/pdf')
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/report/excel', methods=['POST'])
def create_excel_report():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        
        data = request.json
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Analysis"
        
        ws['A1'] = "Real Estate Investment Analysis"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A3'] = "Address:"
        ws['B3'] = data.get('address', '')
        
        if data.get('scenarios'):
            ws['A5'] = "Strategy"
            ws['B5'] = "ROI"
            ws['C5'] = "Profit/Cash Flow"
            for col in ['A5', 'B5', 'C5']:
                ws[col].font = Font(bold=True)
                ws[col].fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
            
            row = 6
            for s in data['scenarios'][:8]:
                ws[f'A{row}'] = s['name']
                ws[f'B{row}'] = f"{s['roi']}%"
                profit = s.get('profit') or s.get('annual_cash_flow', 0)
                ws[f'C{row}'] = profit
                row += 1
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name='analysis_report.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/health', methods=['GET'])
@app.route('/api/health', methods=['GET'])
def health_check():
    return jsonify({'status': 'healthy', 'apify_configured': bool(APIFY_TOKEN)})

@app.route('/')
def index():
    return send_from_directory('../public', 'index.html')

@app.route('/<path:path>')
def static_files(path):
    return send_from_directory('../public', path)

# Vercel handler
app = app
